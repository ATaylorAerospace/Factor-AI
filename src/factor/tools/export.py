"""Report export tools — Excel, HTML, and structured JSON."""

from __future__ import annotations

import json
import logging
from datetime import datetime, timezone
from pathlib import Path

from strands import tool

from factor import DISCLAIMER

logger = logging.getLogger(__name__)


@tool
def build_risk_report(analysis_results: dict) -> dict:
    """Assemble structured risk report from analysis results.

    Labels all synthetic content appropriately.

    Args:
        analysis_results: Dictionary containing risk scores, gaps, comparisons.

    Returns:
        Structured report dictionary ready for export.
    """
    risk_scores = analysis_results.get("risk_scores", [])
    gaps = analysis_results.get("gaps", [])
    comparisons = analysis_results.get("comparisons", [])

    critical = [r for r in risk_scores if r.get("risk_level") == "critical"]
    high = [r for r in risk_scores if r.get("risk_level") == "high"]
    medium = [r for r in risk_scores if r.get("risk_level") == "medium"]

    if critical:
        overall_risk = "critical"
    elif high:
        overall_risk = "high"
    elif medium:
        overall_risk = "medium"
    else:
        overall_risk = "low"

    high_severity_gaps = [g for g in gaps if g.get("severity") in ("high", "critical")]

    report = {
        "title": "Due Diligence Risk Report",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "overall_risk": overall_risk,
        "disclaimer": DISCLAIMER,
        "synthetic_dataset_used": True,
        "executive_summary": (
            f"Analysis identified {len(risk_scores)} provisions across "
            f"{analysis_results.get('document_count', 0)} documents. "
            f"Found {len(critical)} critical, {len(high)} high, and {len(medium)} medium "
            f"risk provisions. Identified {len(gaps)} missing provisions "
            f"({len(high_severity_gaps)} high severity)."
        ),
        "sections": [
            {
                "title": "Risk Assessment",
                "items": risk_scores,
                "synthetic_content": True,
            },
            {
                "title": "Gap Analysis",
                "items": gaps,
                "synthetic_content": True,
            },
            {
                "title": "Cross-Document Comparison",
                "items": comparisons,
                "synthetic_content": True,
            },
        ],
    }

    logger.info("Built risk report: overall_risk=%s", overall_risk)
    return report


@tool
def export_excel(report: dict, output_path: str) -> str:
    """Export report to Excel with disclaimer tab about synthetic dataset.

    Args:
        report: Structured report dictionary from build_risk_report.
        output_path: Path where the Excel file should be saved.

    Returns:
        Path to the generated Excel file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # Disclaimer sheet first
    ws_disc = wb.active
    ws_disc.title = "DISCLAIMER"
    ws_disc["A1"] = "IMPORTANT DISCLAIMER"
    ws_disc["A1"].font = Font(bold=True, size=16, color="FF0000")
    ws_disc["A3"] = DISCLAIMER
    ws_disc["A3"].alignment = Alignment(wrap_text=True)
    ws_disc.column_dimensions["A"].width = 100
    ws_disc["A5"] = "Dataset: Taylor658/synthetic-legal (HuggingFace)"
    ws_disc["A6"] = "ALL citations, statutes, and case references are SYNTHETIC."
    ws_disc["A6"].font = Font(bold=True, color="FF0000")

    # Executive Summary
    ws_summary = wb.create_sheet("Executive Summary")
    ws_summary["A1"] = "Due Diligence Risk Report"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A3"] = f"Overall Risk: {report.get('overall_risk', 'N/A').upper()}"
    risk_color = {
        "critical": "FF0000", "high": "FF6600",
        "medium": "FFCC00", "low": "00CC00",
    }
    color = risk_color.get(report.get("overall_risk", ""), "000000")
    ws_summary["A3"].font = Font(bold=True, size=12, color=color)
    ws_summary["A5"] = report.get("executive_summary", "")
    ws_summary["A5"].alignment = Alignment(wrap_text=True)
    ws_summary.column_dimensions["A"].width = 100
    ws_summary["A7"] = f"Generated: {report.get('generated_at', '')}"

    # Risk scores
    risk_items = []
    for section in report.get("sections", []):
        if section.get("title") == "Risk Assessment":
            risk_items = section.get("items", [])

    if risk_items:
        ws_risk = wb.create_sheet("Risk Scores")
        headers = ["Provision ID", "Risk Level", "Score", "Factors", "Explanation"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws_risk.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

        for row, item in enumerate(risk_items, 2):
            ws_risk.cell(row=row, column=1, value=item.get("provision_id", ""))
            ws_risk.cell(row=row, column=2, value=item.get("risk_level", ""))
            ws_risk.cell(row=row, column=3, value=item.get("score", 0))
            ws_risk.cell(row=row, column=4, value="; ".join(item.get("factors", [])))
            ws_risk.cell(row=row, column=5, value=item.get("explanation", ""))

    # Gaps
    gap_items = []
    for section in report.get("sections", []):
        if section.get("title") == "Gap Analysis":
            gap_items = section.get("items", [])

    if gap_items:
        ws_gaps = wb.create_sheet("Gaps")
        headers = ["Missing Provision", "Severity", "Recommendation"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws_gaps.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

        for row, item in enumerate(gap_items, 2):
            ws_gaps.cell(row=row, column=1, value=item.get("missing_provision", ""))
            ws_gaps.cell(row=row, column=2, value=item.get("severity", ""))
            ws_gaps.cell(row=row, column=3, value=item.get("recommendation", ""))

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))

    logger.info("Exported Excel report to %s", output_path)
    return str(path)


@tool
def export_html(report: dict, output_path: str) -> str:
    """Export report to HTML with disclaimers on every page.

    Args:
        report: Structured report dictionary from build_risk_report.
        output_path: Path where the HTML file should be saved.

    Returns:
        Path to the generated HTML file.
    """
    from jinja2 import Environment, FileSystemLoader, BaseLoader

    template_str = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ report.title }} — Factor</title>
    <style>
        body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; margin: 0; padding: 20px; background: #f5f5f5; }
        .container { max-width: 900px; margin: 0 auto; background: white; padding: 40px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        .disclaimer { background: #fff3cd; border: 2px solid #ffc107; padding: 16px; border-radius: 8px; margin: 20px 0; font-size: 0.9em; }
        .disclaimer strong { color: #856404; }
        h1 { color: #1a1a2e; }
        h2 { color: #16213e; border-bottom: 2px solid #0f3460; padding-bottom: 8px; }
        .risk-badge { display: inline-block; padding: 4px 12px; border-radius: 4px; color: white; font-weight: bold; text-transform: uppercase; }
        .risk-critical { background: #dc3545; }
        .risk-high { background: #fd7e14; }
        .risk-medium { background: #ffc107; color: #333; }
        .risk-low { background: #28a745; }
        table { width: 100%; border-collapse: collapse; margin: 16px 0; }
        th { background: #4472C4; color: white; padding: 10px; text-align: left; }
        td { padding: 10px; border-bottom: 1px solid #ddd; }
        .footer { text-align: center; color: #666; margin-top: 40px; font-size: 0.85em; }
    </style>
</head>
<body>
<div class="container">
    <div class="disclaimer">
        <strong>⚠️ DISCLAIMER:</strong> {{ report.disclaimer }}
    </div>
    <h1>{{ report.title }}</h1>
    <p><strong>Generated:</strong> {{ report.generated_at }}</p>
    <p><strong>Overall Risk:</strong>
        <span class="risk-badge risk-{{ report.overall_risk }}">{{ report.overall_risk }}</span>
    </p>
    <h2>Executive Summary</h2>
    <p>{{ report.executive_summary }}</p>
    {% for section in report.sections %}
    <h2>{{ section.title }}</h2>
    <div class="disclaimer">
        <strong>⚠️ Note:</strong> Content in this section may reference synthetic data from Taylor658/synthetic-legal. All citations are artificially generated.
    </div>
    {% if section.items %}
    <table>
        <thead>
            <tr>
                {% for key in section.items[0].keys() %}
                <th>{{ key | replace('_', ' ') | title }}</th>
                {% endfor %}
            </tr>
        </thead>
        <tbody>
            {% for item in section.items %}
            <tr>
                {% for val in item.values() %}
                <td>{{ val }}</td>
                {% endfor %}
            </tr>
            {% endfor %}
        </tbody>
    </table>
    {% endif %}
    {% endfor %}
    <div class="disclaimer">
        <strong>⚠️ DISCLAIMER:</strong> {{ report.disclaimer }}
    </div>
    <div class="footer">
        <p>Factor — Agentic AI Legal Due Diligence Platform | Author: A Taylor</p>
        <p>Dataset: Taylor658/synthetic-legal (ALL content is synthetic)</p>
    </div>
</div>
</body>
</html>"""

    env = Environment(loader=BaseLoader())
    template = env.from_string(template_str)
    html_content = template.render(report=report)

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(html_content)

    logger.info("Exported HTML report to %s", output_path)
    return str(path)
